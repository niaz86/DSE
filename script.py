# -*- coding: utf-8 -*-
"""DSE"""

import requests
from bs4 import BeautifulSoup
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_FILE = "dse_stocks.xlsx"
JSON_FILE  = "dse_stocks.json"
URL        = "https://www.dsebd.org/latest_share_price_scroll_group.php"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="006B6B")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="F2FAFA")
UP_FILL      = PatternFill("solid", start_color="E6F9EC")
DOWN_FILL    = PatternFill("solid", start_color="FDE8E8")
UP_FONT      = Font(name="Arial", size=10, color="1A7A1A", bold=True)
DOWN_FONT    = Font(name="Arial", size=10, color="CC0000", bold=True)
NORMAL_FONT  = Font(name="Arial", size=10)
BORDER = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── NEW: Day Signal Styles ─────────────────────────────────────────────────────
STRONG_BUY_FILL  = PatternFill("solid", start_color="00B050")   # dark green
BUY_FILL         = PatternFill("solid", start_color="E6F9EC")   # light green
NEUTRAL_FILL     = PatternFill("solid", start_color="FFFF99")   # yellow
SELL_FILL        = PatternFill("solid", start_color="FDE8E8")   # light red
STRONG_SELL_FILL = PatternFill("solid", start_color="FF0000")   # dark red

SIGNAL_FONT_DARK  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SIGNAL_FONT_LIGHT = Font(name="Arial", size=10, bold=True, color="000000")

DATA_COLS  = ["DATE", "#", "TRADING CODE", "GROUP", "LTP*", "HIGH", "LOW",
              "CLOSEP*", "YCP*", "CHANGE", "TRADE", "VALUE (mn)", "VOLUME",
              "DAY SIGNAL", "SCORE", "SIGNAL REASON"]   # ✅ 3 new columns
COL_WIDTHS = [12, 5, 16, 7, 9, 9, 9, 10, 9, 9, 9, 12, 14,
              13, 7, 40]                                 # ✅ 3 new widths

# ── Session ───────────────────────────────────────────────────────────────────
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0"})

# ── Fetch ─────────────────────────────────────────────────────────────────────
def fetch_group(group="A"):
    resp = session.get(URL, params={"group": group}, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    rows = []
    for a_tag in soup.select("a.ab1"):
        tr = a_tag.find_parent("tr")
        if not tr:
            continue
        tds = tr.find_all("td")
        if len(tds) < 11:
            continue
        rows.append({
            "#":            tds[0].get_text(strip=True),
            "TRADING CODE": a_tag.get_text(strip=True),
            "LTP*":         tds[2].get_text(strip=True),
            "HIGH":         tds[3].get_text(strip=True),
            "LOW":          tds[4].get_text(strip=True),
            "CLOSEP*":      tds[5].get_text(strip=True),
            "YCP*":         tds[6].get_text(strip=True),
            "CHANGE":       tds[7].get_text(strip=True),
            "TRADE":        tds[8].get_text(strip=True),
            "VALUE (mn)":   tds[9].get_text(strip=True),
            "VOLUME":       tds[10].get_text(strip=True),
        })
    return rows

# ── NEW: Day Signal Engine ─────────────────────────────────────────────────────
def safe_float(val):
    """Convert string to float, return None if not possible."""
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None

def compute_day_signal(stock):
    """
    Score-based signal using available DSE columns.

    Parameters (from research):
      1. CHANGE sign          → +2 positive / -2 negative
      2. LTP vs YCP           → LTP > YCP = bullish (+1) / bearish (-1)
      3. Close position       → LTP near HIGH = strength (+1) / near LOW = weak (-1)
         formula: (LTP - LOW) / (HIGH - LOW) >= 0.6 → bullish
      4. Volume signal        → VOLUME > 0 and TRADE > 0 → +1 else -1
      5. Price range momentum → HIGH > YCP (+1 breakout) / LOW < YCP (-1 breakdown)

    Score  >=  3  → STRONG BUY  🟢🟢
    Score   =  1,2 → BUY        🟢
    Score   =  0  → NEUTRAL     🟡
    Score  = -1,-2 → SELL       🔴
    Score  <= -3  → STRONG SELL 🔴🔴
    """
    ltp    = safe_float(stock.get("LTP*"))
    high   = safe_float(stock.get("HIGH"))
    low    = safe_float(stock.get("LOW"))
    ycp    = safe_float(stock.get("YCP*"))
    change = safe_float(stock.get("CHANGE"))
    volume = safe_float(stock.get("VOLUME"))
    trade  = safe_float(stock.get("TRADE"))

    score   = 0
    reasons = []

    # ── Parameter 1: CHANGE sign (weight: ±2, strongest signal) ──────────────
    if change is not None:
        if change > 0:
            score += 2
            reasons.append(f"Change +{change} ▲")
        elif change < 0:
            score -= 2
            reasons.append(f"Change {change} ▼")
        else:
            reasons.append("Change=0 →")

    # ── Parameter 2: LTP vs YCP ───────────────────────────────────────────────
    if ltp is not None and ycp is not None and ycp != 0:
        if ltp > ycp:
            score += 1
            reasons.append("LTP>YCP ▲")
        elif ltp < ycp:
            score -= 1
            reasons.append("LTP<YCP ▼")

    # ── Parameter 3: Close position within day range (LTP position) ──────────
    if ltp is not None and high is not None and low is not None:
        day_range = high - low
        if day_range > 0:
            position = (ltp - low) / day_range
            if position >= 0.6:
                score += 1
                reasons.append(f"Closed high({position:.0%}) ▲")
            elif position <= 0.4:
                score -= 1
                reasons.append(f"Closed low({position:.0%}) ▼")
            else:
                reasons.append(f"Mid-range({position:.0%}) →")

    # ── Parameter 4: Volume & Trade activity ──────────────────────────────────
    if volume is not None and trade is not None:
        if volume > 0 and trade > 0:
            score += 1
            reasons.append("Active volume ▲")
        else:
            score -= 1
            reasons.append("No volume ▼")

    # ── Parameter 5: Breakout / Breakdown vs YCP ─────────────────────────────
    if high is not None and low is not None and ycp is not None:
        if high > ycp:
            score += 1
            reasons.append("Broke above YCP ▲")
        if low < ycp:
            score -= 1
            reasons.append("Broke below YCP ▼")

    # ── Map score to signal ───────────────────────────────────────────────────
    if score >= 3:
        label = "STRONG BUY"
    elif score >= 1:
        label = "BUY"
    elif score == 0:
        label = "NEUTRAL"
    elif score >= -2:
        label = "SELL"
    else:
        label = "STRONG SELL"

    return label, score, " | ".join(reasons)

# ── Excel helpers ─────────────────────────────────────────────────────────────
def style_header(ws):
    for col_idx, (col, width) in enumerate(zip(DATA_COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def get_or_create_sheet(wb, name):
    safe = name[:31]
    if safe in wb.sheetnames:
        return wb[safe], False
    ws = wb.create_sheet(title=safe)
    style_header(ws)
    return ws, True

def build_row(today, stock, group):
    signal, score, reason = compute_day_signal(stock)
    return [
        today,
        stock.get("#", ""),
        stock.get("TRADING CODE", ""),
        group,
        stock.get("LTP*", ""),
        stock.get("HIGH", ""),
        stock.get("LOW", ""),
        stock.get("CLOSEP*", ""),
        stock.get("YCP*", ""),
        stock.get("CHANGE", ""),
        stock.get("TRADE", ""),
        stock.get("VALUE (mn)", ""),
        stock.get("VOLUME", ""),
        signal,     # ✅ col 14
        score,      # ✅ col 15
        reason,     # ✅ col 16
    ], signal

def apply_row_style(ws, row_num, signal):
    """Color entire row + apply special font to signal cell based on signal."""
    # Row fill mapping
    fill_map = {
        "STRONG BUY":  STRONG_BUY_FILL,
        "BUY":         BUY_FILL,
        "NEUTRAL":     NEUTRAL_FILL,
        "SELL":        SELL_FILL,
        "STRONG SELL": STRONG_SELL_FILL,
    }
    # Signal cell font (white on dark, black on light)
    font_map = {
        "STRONG BUY":  SIGNAL_FONT_DARK,
        "BUY":         SIGNAL_FONT_LIGHT,
        "NEUTRAL":     SIGNAL_FONT_LIGHT,
        "SELL":        SIGNAL_FONT_LIGHT,
        "STRONG SELL": SIGNAL_FONT_DARK,
    }
    row_fill  = fill_map.get(signal, ALT_FILL)
    sig_font  = font_map.get(signal, NORMAL_FONT)
    sig_col   = 14  # DAY SIGNAL column index

    for col_idx in range(1, len(DATA_COLS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill   = row_fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == sig_col:
            cell.font = sig_font
        else:
            cell.font = NORMAL_FONT

# ── JSON helpers ──────────────────────────────────────────────────────────────
def load_json():
    if os.path.isfile(JSON_FILE):
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_json(data):
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def build_json_row(today, stock, group, signal, score, reason):
    return {
        "DATE":         today,
        "#":            stock.get("#", ""),
        "TRADING CODE": stock.get("TRADING CODE", ""),
        "GROUP":        group,
        "LTP*":         stock.get("LTP*", ""),
        "HIGH":         stock.get("HIGH", ""),
        "LOW":          stock.get("LOW", ""),
        "CLOSEP*":      stock.get("CLOSEP*", ""),
        "YCP*":         stock.get("YCP*", ""),
        "CHANGE":       stock.get("CHANGE", ""),
        "TRADE":        stock.get("TRADE", ""),
        "VALUE (mn)":   stock.get("VALUE (mn)", ""),
        "VOLUME":       stock.get("VOLUME", ""),
        "DAY SIGNAL":   signal,    # ✅ NEW
        "SCORE":        score,     # ✅ NEW
        "SIGNAL REASON": reason,   # ✅ NEW
    }

def json_is_duplicate(stock_list, new_row):
    if not stock_list:
        return False
    last = stock_list[-1]
    keys = ["GROUP","LTP*","HIGH","LOW","CLOSEP*","YCP*","CHANGE",
            "TRADE","VALUE (mn)","VOLUME"]
    return all(str(last.get(k,"")) == str(new_row.get(k,"")) for k in keys)

# ── Main ──────────────────────────────────────────────────────────────────────
def run(groups=("A","B","G","N","Z")):

    wb = load_workbook(EXCEL_FILE) if os.path.isfile(EXCEL_FILE) else Workbook()
    json_data = load_json()

    json_created = json_appended = json_duplicate = 0
    signal_counts = {"STRONG BUY": 0, "BUY": 0, "NEUTRAL": 0,
                     "SELL": 0, "STRONG SELL": 0}

    today = datetime.now().strftime("%Y-%m-%d")

    for group in groups:
        stocks = fetch_group(group)

        for stock in stocks:
            code = stock.get("TRADING CODE", "").strip()
            if not code:
                continue

            # ── Excel ────────────────────────────────────────────────────────
            ws, is_new = get_or_create_sheet(wb, code)
            row_data, signal = build_row(today, stock, group)
            ws.append(row_data)
            current_row = ws.max_row
            apply_row_style(ws, current_row, signal)   # ✅ color entire row

            signal_counts[signal] = signal_counts.get(signal, 0) + 1

            # ── JSON ─────────────────────────────────────────────────────────
            _, score, reason = compute_day_signal(stock)
            json_row = build_json_row(today, stock, group, signal, score, reason)

            if code not in json_data:
                json_data[code] = [json_row]
                json_created += 1
            else:
                if json_is_duplicate(json_data[code], json_row):
                    json_duplicate += 1
                else:
                    json_data[code].append(json_row)
                    json_appended += 1

    wb.save(EXCEL_FILE)
    save_json(json_data)

    print("\n=== SIGNAL SUMMARY ===")
    for sig, count in signal_counts.items():
        print(f"  {sig:12s}: {count}")

    print("\n=== JSON SUMMARY ===")
    print(f"  New       : {json_created}")
    print(f"  Appended  : {json_appended}")
    print(f"  Duplicate : {json_duplicate}")

if __name__ == "__main__":
    run()

with open("last_run.txt", "w") as f:
    f.write(datetime.now().strftime("%Y-%m-%d"))
