import requests
import json
import time
import os
import threading
from datetime import date
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_URL      = "https://www.dsebd.org"
GROUP_URL     = BASE_URL + "/latest_share_price_scroll_group.php"
COMPANY_URL   = BASE_URL + "/displayCompany.php"
CATEGORIES    = ["A", "B", "G", "N", "Z"]
OUTPUT_JSON   = os.environ.get("DSE_OUTPUT_JSON",  "dse_companies.json")
OUTPUT_EXCEL  = os.environ.get("DSE_OUTPUT_EXCEL", "dse_companies.xlsx")
REQUEST_DELAY = 1.5   # reduced — parallelism spreads load naturally
MAX_WORKERS   = 10    # tune this up/down based on your connection & DSE tolerance

HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/146.0.0.0 Safari/537.36"
    ),
    "Referer": BASE_URL,
}

# ── Column layout for each stock sheet ────────────────────────────────────────
DAILY_COLUMNS = [
    ("Date",                    "last_updated"),
    ("Opening Price",           "opening_price"),
    ("Market Cap (mn)",         "market_cap_mn"),
    ("52W High",                "week52_high"),
    ("52W Low",                 "week52_low"),
    ("Trailing P/E",            "trailing_pe"),
    ("EPS Basic",               "latest_eps_basic"),
    ("NAV Per Share",           "nav_per_share"),
    ("Cash Dividend",           "cash_dividend"),
    ("Bonus Dividend",          "bonus_stock_dividend"),
    ("Paid-up Capital (mn)",    "paid_up_capital_mn"),
    ("Authorized Capital (mn)", "authorized_capital_mn"),
    ("Sponsor/Dir %",           "sponsor_director"),
    ("Institute %",             "institute"),
    ("Public %",                "public"),
]


# ── Scraping helpers (unchanged) ──────────────────────────────────────────────
def fetch_page(trading_code):
    resp = requests.get(COMPANY_URL, params={"name": trading_code}, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")

def find_h2(soup, text):
    for h2 in soup.find_all("h2"):
        if text in h2.get_text():
            return h2
    return None

def find_by_label(soup, label):
    for th in soup.find_all("th"):
        if label.lower() in th.get_text(strip=True).lower():
            td = th.find_next_sibling("td")
            if td:
                return td.get_text(" ", strip=True)
    for td in soup.find_all("td"):
        if td.get_text(strip=True).lower() == label.lower():
            sibling = td.find_next_sibling("td")
            if sibling:
                return sibling.get_text(" ", strip=True)
    return ""

def get_company_name(soup):
    header = soup.find("h2", class_="BodyHead")
    if not header:
        return ""
    full_text = header.get_text(" ", strip=True)
    return full_text.split("Company Name:")[-1].strip() if "Company Name:" in full_text else full_text

def get_market_info(soup):
    week52_range = find_by_label(soup, "52 Weeks")
    week52_high = week52_low = ""
    if " - " in week52_range:
        parts = week52_range.split(" - ")
        week52_low, week52_high = parts[0].strip(), parts[1].strip()
    return {
        "opening_price": find_by_label(soup, "Opening Price"),
        "market_cap_mn": find_by_label(soup, "Market Capitalization"),
        "week52_high":   week52_high,
        "week52_low":    week52_low,
    }

def get_basic_info(soup):
    return {
        "authorized_capital_mn": find_by_label(soup, "Authorized Capital (mn)"),
        "paid_up_capital_mn":    find_by_label(soup, "Paid-up Capital (mn)"),
        "face_par_value":        find_by_label(soup, "Face/par Value"),
        "sector":                find_by_label(soup, "Sector"),
        "debut_trading_date":    find_by_label(soup, "Debut Trading Date"),
    }

def get_latest_shareholding(soup):
    result = {}
    rows = [td.parent for td in soup.find_all("td") if "Share Holding Percentage" in td.get_text()]
    if rows:
        inner_cells = rows[-1].find_all("td", style=lambda v: v and "border:hidden" in v)
        for label, cell in zip(["sponsor_director", "govt", "institute", "foreign", "public"], inner_cells):
            text = cell.get_text(" ", strip=True)
            result[label] = text.split(":")[-1].strip() if ":" in text else text
    return result

def get_latest_eps(soup):
    header = find_h2(soup, "Interim Financial Performance")
    if not header:
        return ""
    table = header.find_next("table")
    if not table:
        return ""
    in_continuing = False
    for row in table.find_all("tr"):
        cells = row.find_all(["th", "td"])
        if not cells:
            continue
        first = cells[0].get_text(strip=True)
        if "continuing" in first.lower() and "EPS" in first:
            in_continuing = True
            continue
        if in_continuing and first == "Basic":
            for v in reversed([c.get_text(strip=True) for c in cells[1:-1]]):
                if v and v != "-":
                    return v
            break
    return ""

def get_latest_nav(soup):
    header = find_h2(soup, "Financial Performance as per Audited")
    if not header:
        return ""
    table = header.find_next("table")
    if not table:
        return ""
    data_rows = table.find_all("tr", class_=lambda c: c and "shrink" in c)
    if not data_rows:
        return ""
    cells = data_rows[-1].find_all(["th", "td"])
    return cells[7].get_text(strip=True) if len(cells) > 7 else ""

def get_dividends(soup):
    return {
        "cash_dividend":        find_by_label(soup, "Cash Dividend"),
        "bonus_stock_dividend": find_by_label(soup, "Bonus Issue (Stock Dividend)"),
    }

def get_trailing_pe(soup):
    header = find_h2(soup, "P/E Ratio based on latest Un-audited")
    if not header:
        header = find_h2(soup, "latest Un-audited Financial Statements")
    if not header:
        return ""
    table = header.find_next("table")
    if not table:
        return ""
    for row in table.find_all("tr"):
        cells = row.find_all(["th", "td"])
        if cells and "Trailing P/E" in cells[0].get_text():
            for v in reversed([c.get_text(strip=True) for c in cells[1:]]):
                if v and v != "-":
                    return v
    return ""

def get_other_info(soup):
    return {
        "listing_year":    find_by_label(soup, "Listing Year"),
        "market_category": find_by_label(soup, "Market Category"),
    }

def get_address(soup):
    head_office = contact_phone = email = web_address = ""
    for td in soup.find_all("td"):
        text = td.get_text(strip=True)
        if text == "Head Office":
            s = td.find_next_sibling("td")
            if s: head_office = s.get_text(strip=True)
        elif text == "Contact Phone":
            s = td.find_next_sibling("td")
            if s: contact_phone = s.get_text(strip=True)
        elif text == "Web Address":
            s = td.find_next_sibling("td")
            if s: web_address = s.get_text(strip=True)
    for td in soup.find_all("td"):
        if td.get_text(strip=True) == "E-mail":
            s = td.find_next_sibling("td")
            if s:
                email = s.get_text(strip=True)
                break
    return {"head_office": head_office, "contact_phone": contact_phone,
            "email": email, "web_address": web_address}

def scrape_company(trading_code):
    soup       = fetch_page(trading_code)
    market     = get_market_info(soup)
    basic      = get_basic_info(soup)
    other      = get_other_info(soup)
    dividends  = get_dividends(soup)
    holding    = get_latest_shareholding(soup)

    return {
        "trading_code":          trading_code,
        "company_name":          get_company_name(soup),
        "sector":                basic["sector"],
        "face_par_value":        basic["face_par_value"],
        "debut_trading_date":    basic["debut_trading_date"],
        "listing_year":          other["listing_year"],
        "market_category":       other["market_category"],
        "address":               get_address(soup),
        "authorized_capital_mn": basic["authorized_capital_mn"],
        "paid_up_capital_mn":    basic["paid_up_capital_mn"],
        "nav_per_share":         get_latest_nav(soup),
        "latest_eps_basic":      get_latest_eps(soup),
        "cash_dividend":         dividends["cash_dividend"],
        "bonus_stock_dividend":  dividends["bonus_stock_dividend"],
        "sponsor_director":      holding.get("sponsor_director", ""),
        "institute":             holding.get("institute", ""),
        "public":                holding.get("public", ""),
        "week52_high":           market["week52_high"],
        "week52_low":            market["week52_low"],
        "trailing_pe":           get_trailing_pe(soup),
        "opening_price":         market["opening_price"],
        "market_cap_mn":         market["market_cap_mn"],
        "last_updated":          str(date.today()),
    }


# ── Excel writer (unchanged — runs single-threaded after all scraping) ─────────
def safe_sheet_name(code):
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        code = code.replace(ch, "_")
    return code[:31]

def style_header_cell(cell, text):
    cell.value = cell.value if text is None else text
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="004B5A")
    cell.alignment = Alignment(horizontal="center")

def write_static_block(ws, data):
    static_fields = [
        ("Company Name",    data["company_name"]),
        ("Trading Code",    data["trading_code"]),
        ("Sector",          data["sector"]),
        ("Market Category", data["market_category"]),
        ("Face Value",      data["face_par_value"]),
        ("Listing Year",    data["listing_year"]),
        ("Debut Date",      data["debut_trading_date"]),
        ("Address",         data["address"]["head_office"]),
        ("Phone",           data["address"]["contact_phone"]),
        ("Email",           data["address"]["email"]),
    ]
    for i, (label, value) in enumerate(static_fields, start=1):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    for col_idx, (header_label, _) in enumerate(DAILY_COLUMNS, start=1):
        cell = ws.cell(row=12, column=col_idx, value=header_label)
        style_header_cell(cell, header_label)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    for col_idx in range(1, len(DAILY_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

def append_daily_row(ws, data):
    next_row = 13
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1
    for col_idx, (_, key) in enumerate(DAILY_COLUMNS, start=1):
        ws.cell(row=next_row, column=col_idx, value=data.get(key, ""))

def update_excel(filepath, all_data):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for data in all_data:
        sheet_name = safe_sheet_name(data["trading_code"])
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            write_static_block(ws, data)
        append_daily_row(ws, data)

    wb.save(filepath)
    print(f"Excel workbook saved: '{filepath}' ({len(wb.sheetnames)} sheets).")


# ── Discovery — parallel across categories ────────────────────────────────────
def get_trading_codes(category):
    print(f"  Discovering codes for category {category}...")
    resp = requests.get(GROUP_URL, params={"group": category}, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    codes = []
    for a in soup.find_all("a", href=True):
        if a["href"].startswith("displayCompany.php?name="):
            code = a["href"].split("=")[-1].strip()
            if code:
                codes.append(code)
    print(f"    Found {len(codes)} codes in category {category}.")
    return codes

def discover_all_codes():
    """Fetch all category pages in parallel, then deduplicate."""
    all_codes = []
    with ThreadPoolExecutor(max_workers=len(CATEGORIES)) as ex:
        futures = {ex.submit(get_trading_codes, cat): cat for cat in CATEGORIES}
        for future in as_completed(futures):
            try:
                all_codes.extend(future.result())
            except Exception as e:
                print(f"  Category {futures[future]} discovery failed: {e}")

    seen, unique = set(), []
    for code in all_codes:
        if code not in seen:
            seen.add(code)
            unique.append(code)
    print(f"\nTotal unique trading codes discovered: {len(unique)}\n")
    return unique


# ── Worker: scrape one company with a small per-thread delay ──────────────────
def scrape_worker(args):
    """
    Wraps scrape_company for use in the thread pool.
    Returns (index, code, data_or_None, error_or_None).
    The small sleep staggers requests so we don't hammer the server
    with all MAX_WORKERS threads firing simultaneously at t=0.
    """
    index, total, code = args
    time.sleep(REQUEST_DELAY)
    try:
        data = scrape_company(code)
        return index, total, code, data, None
    except Exception as e:
        return index, total, code, None, e


# ── Main orchestrator ─────────────────────────────────────────────────────────
def run_daily_scrape():
    print(f"=== DSE Daily Scrape — {date.today()} ===\n")
    print(f"Thread pool size: {MAX_WORKERS}\n")

    all_codes = discover_all_codes()
    total     = len(all_codes)

    store   = {}
    scraped = []           # ordered list for Excel (filled after pool finishes)
    lock    = threading.Lock()
    success = fail = 0

    # Build argument tuples so each worker knows its position for logging
    work_items = [(i, total, code) for i, code in enumerate(all_codes, 1)]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_worker, item): item for item in work_items}

        for future in as_completed(futures):
            index, total_, code, data, err = future.result()
            if err:
                with lock:
                    fail += 1
                print(f"[{index}/{total_}] {code} — FAILED: {err}")
            else:
                with lock:
                    success += 1
                    store[code] = data
                print(f"[{index}/{total_}] {code} — OK")

    # Preserve the original discovery order for stable Excel sheet ordering
    scraped = [store[code] for code in all_codes if code in store]

    # ── Save outputs (single-threaded) ────────────────────────────────────────
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(store, f, indent=2, ensure_ascii=False)
    print(f"\nJSON backup saved: '{OUTPUT_JSON}'.")

    update_excel(OUTPUT_EXCEL, scraped)
    print(f"\nDone. {success} succeeded, {fail} failed.")


if __name__ == "__main__":
    run_daily_scrape()
