import requests
from bs4 import BeautifulSoup
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── CONFIG ─────────────────────────────────────────────
FILE_NAME = "dse_market_summary.xlsx"
URL = "https://www.dsebd.org/market_summary.php"

today = datetime.now().strftime("%Y-%m-%d")

params = {
    "startDate": today,
    "endDate": today,
    "archive": "data"
}

headers = {
    "User-Agent": "Mozilla/5.0"
}

# ── FETCH DATA ─────────────────────────────────────────
res = requests.get(URL, params=params, headers=headers)
soup = BeautifulSoup(res.text, "html.parser")

tables = soup.select("table.table")

data_list = []

for table in tables:
    rows = table.find_all("tr")

    if len(rows) < 2:
        continue

    title = rows[0].get_text(strip=True)
    if "Market Summary of" not in title:
        continue

    date = title.replace("Market Summary of", "").strip()

    data = {"Date": date}

    for row in rows[1:]:
        cols = row.find_all("td")
        if len(cols) == 4:
            data[cols[0].get_text(strip=True)] = cols[1].get_text(strip=True)
            data[cols[2].get_text(strip=True)] = cols[3].get_text(strip=True)

    data_list.append(data)

# ── GET ALL COLUMNS ────────────────────────────────────
all_keys = set()
for d in data_list:
    all_keys.update(d.keys())

columns = list(all_keys)

# ── STYLES ─────────────────────────────────────────────
header_fill = PatternFill("solid", start_color="006B6B")
header_font = Font(color="FFFFFF", bold=True)

good_fill = PatternFill("solid", start_color="C6EFCE")   # green
bad_fill  = PatternFill("solid", start_color="FFC7CE")   # red
normal_fill = PatternFill("solid", start_color="FFFFFF") # white

# ── LOAD OR CREATE EXCEL ───────────────────────────────
if os.path.exists(FILE_NAME):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    columns = [cell.value for cell in ws[1]]
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Summary"

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

# ── CHECK EXISTING DATES ───────────────────────────────
existing_dates = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    if row:
        existing_dates.add(row[0])

# ── APPEND DATA ────────────────────────────────────────
for d in data_list:
    if d.get("Date") in existing_dates:
        print(f"⏩ Skipped (already exists): {d.get('Date')}")
        continue

    row_idx = ws.max_row + 1

    for col_idx, col_name in enumerate(columns, 1):
        value = d.get(col_name, "")
        ws.cell(row=row_idx, column=col_idx, value=value).alignment = Alignment(horizontal="center")

    # 🎨 Color logic
    change = d.get("DSEX Index Change", "0")

    try:
        change_val = float(str(change).replace(",", ""))
    except:
        change_val = 0

    if change_val > 0:
        fill = good_fill
    elif change_val < 0:
        fill = bad_fill
    else:
        fill = normal_fill

    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

    print(f"✅ Added: {d.get('Date')}")

# ── SAVE FILE ──────────────────────────────────────────
wb.save(FILE_NAME)

print("📊 Excel updated successfully!")